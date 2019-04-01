#Python 2.7
import warnings
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Font

warnings.simplefilter("ignore")

filename = raw_input("Enter the file name:  ")		#OPENING SPREADSHEET
filename = filename + ".xlsx"
try:
	wbr = openpyxl.load_workbook(filename.strip())
except:
	print "\nError: \"" + filename + "\" file not found."
	exit()

sheetname = raw_input("Enter the worksheet name: ")		#OPENING WORKSHEET
try:
	wsr = wbr[sheetname.strip()]
except:
	print "\nError: \"" + sheetname + "\" worksheet does not exist."
	exit()

row = -1		#TRACKING HEADER ROW
for i in range(1, 10):
	if wsr.cell(row = i, column = 1).value.strip() == "Issue" or wsr.cell(row = i, column = 1).value.strip() == "Issues":
		row = i
		break
if row == -1: 
	print "\nError: Header row could not be found."
	exit()

col = -1		#TRACKING MODULE COLUMN
for i in range(1, 15):		
	if wsr.cell(row = row, column = i).value.strip() == "Module" or wsr.cell(row = row, column = i).value.strip() == "Modules": 
		col = i
		break
if col == -1: 
	print "\nError: \"Module\" column could not be found."
	exit()

modules = dict()
flag = False
first = True

for cell in wsr.columns[col - 1]:			#CALCULATION
	if first:
		first =  False
		continue
	flag = False
	if cell.value is None: continue
	for module in modules:
		if cell.value.strip().lower() == module.lower():
			flag = True			
			modules[module] += 1
			break
	if flag == False:	
		modules[cell.value.strip()] = 1

for module in modules:		#OUTPUT ON CONSOLE
	print module + ":", modules[module]

wbw = openpyxl.Workbook()		
wsw = wbw.active
wsw.title = "Sheet1"

wsw['A1'].alignment = Alignment(horizontal = 'center')		#FORMATTING
wsw['B1'].alignment = Alignment(horizontal = 'center')
wsw['A1'].font = Font(name = "Arial", size = 11, bold = True)
wsw['B1'].font = Font(name = "Arial", size = 11, bold = True)
wsw.column_dimensions['A'].width = 40
wsw.column_dimensions['B'].width = 15

wsw['A1'] = "Modules"		#OUTPUT DATA ON EXCEL SHEET
wsw['B1'] = "Occurrences"
for module in modules:
	wsw.append([module, modules[module]])

wbw.save("Report.xlsx")