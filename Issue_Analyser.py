#Python 2.7
import warnings
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.chart.label import DataLabelList

warnings.simplefilter("ignore")

filename = raw_input("Enter the file name:  ")		#OPENING SPREADSHEET
filename = filename + ".xlsx"
try:
	wbr = openpyxl.load_workbook(filename.strip())
except:
	print "\nError: \"" + filename + "\" file not found."
	exit()

sheetname = raw_input("Enter the worksheet name: ")		#OPENING WORKSHEET
try:
	wsr = wbr[sheetname.strip()]
except:
	print "\nError: \"" + sheetname + "\" worksheet does not exist."
	exit()

row = -1		#TRACKING HEADER ROW
for i in range(1, 10):
	if wsr.cell(row = i, column = 1).value.strip() == "Issue" or "Issues":
		row = i
		break
if row == -1: 
	print "\nError: Header row could not be found."
	exit()

col = -1		#TRACKING STATUS COLUMN
for i in range(1, 15):		
	if wsr.cell(row = row, column = i).value.strip() == "Status": 
		col = i
		break
if col == -1: 
	print "\nError: \"Status\" column could not be found."
	exit()
 
issues = openn = closed = feedback = acknowledged = 0
 
for cell in wsr.columns[col - 1]:			#CALCULATION
	if cell.value is None: break
	elif cell.value == "Open": openn += 1
	elif cell.value == "Closed": closed += 1
	elif cell.value == "Feedback": feedback += 1
	elif cell.value == "Acknowledged": acknowledged += 1
	issues += 1
issues -= 1
	
print "Total number of issues:", issues		#OUTPUT ON CONSOLE
print "Total issues open:", openn
print "Total issues closed:", closed
print "Total issues in feedback:", feedback
print "Total issues acknowledged:", acknowledged

wbw = openpyxl.Workbook()		
wsw = wbw.active
wsw.title = "Sheet1"

for i in range(1, 3):		#FORMATTING
	for j in range(1, 7):
		wsw.cell(row = j, column = i).font = Font(name = "Arial", size = 12, bold = True)
for i in range(2, 7):
	wsw.cell(row = i, column = 2).alignment = Alignment(horizontal = 'center')
wsw.column_dimensions['A'].width = 30
wsw.merge_cells('A1:B1')
wsw['A1'].alignment = Alignment(horizontal = 'center')

wsw['A1'] = "Bug Analysis"		#OUTPUT DATA ON EXCEL SHEET
wsw['A2'] = "Total number of issues"		
wsw['B2'] = issues
wsw['A3'] = "Total issues open"
wsw['B3'] = openn
wsw['A4'] = "Total issues closed"
wsw['B4'] = closed
wsw['A5'] = "Total issues in feedback"
wsw['B5'] = feedback
wsw['A6'] = "Total issues acknowledged"
wsw['B6'] = acknowledged

data = openpyxl.chart.Reference(wsw, min_col = 2, min_row = 3, max_col = 2, max_row = 6)			#OUTPUT PIE CHART ON EXCEL SHEET
categories = openpyxl.chart.Reference(wsw, min_col = 1, min_row = 3, max_col = 1, max_row = 6)
chartobj = openpyxl.chart.PieChart()
chartobj.add_data(data)
chartobj.set_categories(categories)
chartobj.dataLabels = DataLabelList()
chartobj.dataLabels.showVal = True
wsw.add_chart(chartobj, 'A8')

wbw.save('Report.xlsx')